import shutil
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from app.utils.excel_writer import FIELDS_HEADERS, TABLES_HEADERS, write_row, write_rows
from app.models import InputData, ModifyTableInput, NewField

SAMPLE_SQL = (
    "CREATE TABLE `ai_media_task` (\n"
    "  `id` bigint(20) unsigned NOT NULL AUTO_INCREMENT COMMENT '自增主键ID'\n"
    ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='AI媒体任务表'"
)

SAMPLE_DATA = InputData(
    mysql_sql=SAMPLE_SQL,
    day_or_hour="天表",
    product_line="sfst",
)


class TestWriteRow(unittest.TestCase):
    def setUp(self):
        self.tmp_dir = Path(tempfile.mkdtemp())
        self.excel_path = self.tmp_dir / "test_output.xlsx"

    def tearDown(self):
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def test_creates_file_and_writes_row(self):
        write_row(self.excel_path, SAMPLE_DATA)
        self.assertTrue(self.excel_path.exists())

        wb = load_workbook(self.excel_path)
        tables_ws = wb["tables"]
        self.assertEqual(tables_ws.max_row, 2)

        row = list(tables_ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(row[0], "ai_media_task")
        self.assertEqual(row[1], "sfst")
        self.assertEqual(row[2], "天表")
        self.assertIsNone(row[3])

    def test_correct_sheets_and_headers(self):
        write_row(self.excel_path, SAMPLE_DATA)
        wb = load_workbook(self.excel_path)
        self.assertIn("tables", wb.sheetnames)
        self.assertIn("fields", wb.sheetnames)
        self.assertEqual(
            [cell.value for cell in wb["tables"][1]], TABLES_HEADERS
        )
        self.assertEqual(
            [cell.value for cell in wb["fields"][1]], FIELDS_HEADERS
        )

    def test_fields_sheet(self):
        write_row(self.excel_path, SAMPLE_DATA)
        wb = load_workbook(self.excel_path)
        fields_ws = wb["fields"]
        self.assertEqual(fields_ws.max_row, 2)

        row = list(fields_ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(row[0], "ai_media_task")
        self.assertEqual(row[5], SAMPLE_SQL)

    def test_overwrite_on_rerun(self):
        write_row(self.excel_path, SAMPLE_DATA)

        second_data = InputData(
            mysql_sql="CREATE TABLE `other_table` (id int);",
            day_or_hour="小时表",
            product_line="wax",
        )
        write_row(self.excel_path, second_data)

        wb = load_workbook(self.excel_path)
        self.assertEqual(wb["tables"].max_row, 2)
        self.assertEqual(wb["fields"].max_row, 2)

        row = list(wb["tables"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(row[0], "other_table")
        self.assertEqual(row[1], "wax")
        self.assertEqual(row[2], "小时表")

    def test_invalid_sql_skips_write(self):
        bad_data = InputData(
            mysql_sql="SELECT * FROM foo",
            day_or_hour="天表",
            product_line="sfst",
        )
        write_row(self.excel_path, bad_data)
        self.assertFalse(self.excel_path.exists())

    def test_sql_preserved_exactly(self):
        write_row(self.excel_path, SAMPLE_DATA)
        wb = load_workbook(self.excel_path)
        fields_ws = wb["fields"]
        row = list(fields_ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(row[5], SAMPLE_SQL)

    def test_with_all_optional_fields(self):
        data = InputData(
            mysql_sql=SAMPLE_SQL,
            day_or_hour="天表",
            product_line="sfst",
            dw_layer="ods",
            table_format="orc",
            target_table_format="hive",
            operate_type="新建表",
        )
        write_row(self.excel_path, data)
        wb = load_workbook(self.excel_path)
        tables_row = list(wb["tables"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
        fields_row = list(wb["fields"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(tables_row[4], "ods")
        self.assertEqual(tables_row[5], "orc")
        self.assertEqual(tables_row[6], "hive")
        self.assertEqual(tables_row[7], "新建表")
        self.assertEqual(fields_row[4], "新建表")

    def test_operate_type_consistency(self):
        data = InputData(
            mysql_sql=SAMPLE_SQL,
            day_or_hour="天表",
            product_line="sfst",
            operate_type="修改表",
        )
        write_row(self.excel_path, data)
        wb = load_workbook(self.excel_path)
        tables_op = list(wb["tables"].iter_rows(min_row=2, max_row=2, values_only=True))[0][7]
        fields_op = list(wb["fields"].iter_rows(min_row=2, max_row=2, values_only=True))[0][4]
        self.assertEqual(tables_op, fields_op, "tables 与 fields 的操作类型必须一致")

    def test_optional_fields_empty(self):
        write_row(self.excel_path, SAMPLE_DATA)
        wb = load_workbook(self.excel_path)
        tables_row = list(wb["tables"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
        fields_row = list(wb["fields"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertIsNone(tables_row[4])
        self.assertIsNone(tables_row[5])
        self.assertIsNone(tables_row[6])
        self.assertIsNone(tables_row[7])
        self.assertIsNone(fields_row[4])
        self.assertEqual(tables_row[0], "ai_media_task")
        self.assertEqual(tables_row[1], "sfst")

    def test_partial_optional_fields(self):
        data = InputData(
            mysql_sql=SAMPLE_SQL,
            day_or_hour="天表",
            product_line="sfst",
            dw_layer="mds",
            operate_type="修改表",
        )
        write_row(self.excel_path, data)
        wb = load_workbook(self.excel_path)
        tables_row = list(wb["tables"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
        fields_row = list(wb["fields"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(tables_row[4], "mds")
        self.assertIsNone(tables_row[5])
        self.assertIsNone(tables_row[6])
        self.assertEqual(tables_row[7], "修改表")
        self.assertEqual(fields_row[4], "修改表")

    def test_creates_parent_directory(self):
        nested_path = self.tmp_dir / "sub" / "dir" / "output.xlsx"
        write_row(nested_path, SAMPLE_DATA)
        self.assertTrue(nested_path.exists())

    def test_table_comment_written_to_tables_sheet(self):
        data = InputData(
            mysql_sql=SAMPLE_SQL,
            day_or_hour="天表",
            product_line="sfst",
            table_comment="AI媒体任务表",
        )
        write_row(self.excel_path, data)
        wb = load_workbook(self.excel_path)
        tables_row = list(wb["tables"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(tables_row[3], "AI媒体任务表")

    def test_table_comment_none_leaves_cell_empty(self):
        data = InputData(
            mysql_sql=SAMPLE_SQL,
            day_or_hour="天表",
            product_line="sfst",
            table_comment=None,
        )
        write_row(self.excel_path, data)
        wb = load_workbook(self.excel_path)
        tables_row = list(wb["tables"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertIsNone(tables_row[3])

    def test_is_sharding_yes_written(self):
        data = InputData(
            mysql_sql=SAMPLE_SQL,
            day_or_hour="天表",
            product_line="sfst",
            is_sharding="是",
        )
        write_row(self.excel_path, data)
        wb = load_workbook(self.excel_path)
        tables_row = list(wb["tables"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(tables_row[9], "是")

    def test_sharding_strips_table_name_suffix(self):
        sharding_sql = "CREATE TABLE `order_0` (`id` int) ENGINE=InnoDB"
        data = InputData(
            mysql_sql=sharding_sql,
            day_or_hour="天表",
            product_line="sfst",
            is_sharding="是",
        )
        write_row(self.excel_path, data)
        wb = load_workbook(self.excel_path)
        tables_row = list(wb["tables"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
        fields_row = list(wb["fields"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(tables_row[0], "order")
        self.assertEqual(fields_row[0], "order")

    def test_is_sharding_yes_no_suffix_unchanged(self):
        data = InputData(
            mysql_sql=SAMPLE_SQL,
            day_or_hour="天表",
            product_line="sfst",
            is_sharding="是",
        )
        write_row(self.excel_path, data)
        wb = load_workbook(self.excel_path)
        tables_row = list(wb["tables"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(tables_row[0], "ai_media_task")

    def test_is_sharding_default_no(self):
        write_row(self.excel_path, SAMPLE_DATA)
        wb = load_workbook(self.excel_path)
        tables_row = list(wb["tables"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(tables_row[9], "否")

    def test_write_rows_multiple_tables(self):
        """write_rows 应批量写入多张表到同一 Excel。"""
        second_data = InputData(
            mysql_sql="CREATE TABLE `other_table` (id int);",
            day_or_hour="小时表",
            product_line="wax",
        )
        write_rows(self.excel_path, [SAMPLE_DATA, second_data])

        wb = load_workbook(self.excel_path)
        tables_ws = wb["tables"]
        fields_ws = wb["fields"]
        self.assertEqual(tables_ws.max_row, 3)
        self.assertEqual(fields_ws.max_row, 3)

        row1 = list(tables_ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        row2 = list(tables_ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        self.assertEqual(row1[0], "ai_media_task")
        self.assertEqual(row2[0], "other_table")

    def test_write_rows_modify_table_multiple_field_rows(self):
        mod = ModifyTableInput(
            table_name="ods_ad_sfst_x_day",
            target_table_format="hive",
            operate_type="修改表",
            new_fields=[
                NewField("c1", "bigint"),
                NewField("c2", "string"),
            ],
        )
        write_rows(self.excel_path, [mod])

        wb = load_workbook(self.excel_path)
        self.assertEqual(wb["tables"].max_row, 2)
        self.assertEqual(wb["fields"].max_row, 3)

        trow = list(wb["tables"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(trow[0], "ods_ad_sfst_x_day")
        self.assertEqual(trow[6], "hive")
        self.assertEqual(trow[7], "修改表")
        self.assertEqual(trow[8], "ods_ad_sfst_x_day")

        f1 = list(wb["fields"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
        f2 = list(wb["fields"].iter_rows(min_row=3, max_row=3, values_only=True))[0]
        self.assertEqual(f1[0], "ods_ad_sfst_x_day")
        self.assertEqual(f1[1], "c1")
        self.assertEqual(f1[2], "bigint")
        self.assertEqual(f1[4], "修改表")
        self.assertEqual(f2[1], "c2")


if __name__ == "__main__":
    unittest.main()
