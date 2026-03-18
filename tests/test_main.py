import json
import shutil
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

SAMPLE_SQL = (
    "CREATE TABLE `ai_media_task` (\n"
    "  `id` bigint(20) unsigned NOT NULL AUTO_INCREMENT COMMENT '自增主键ID'\n"
    ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='AI媒体任务表'"
)

SAMPLE_JSON_DICT = {
    "mysql_sql": SAMPLE_SQL,
    "day_or_hour": "天表",
    "product_line": "sfst",
}

SAMPLE_JSON_STR = json.dumps(SAMPLE_JSON_DICT, ensure_ascii=False)


class TestMainWithJson(unittest.TestCase):
    """通过 --json 参数传入 JSON 字符串的主流程集成测试。"""

    def setUp(self):
        self.tmp_dir = Path(tempfile.mkdtemp())
        self.excel_path = self.tmp_dir / "output.xlsx"

    def tearDown(self):
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def test_json_arg_creates_excel(self):
        with patch(
            "sys.argv",
            ["prog", "--json", SAMPLE_JSON_STR, "--excel", str(self.excel_path)],
        ):
            from app.main import main
            main()

        self.assertTrue(self.excel_path.exists())
        wb = load_workbook(self.excel_path)
        tables_ws = wb["tables"]
        self.assertEqual(tables_ws.max_row, 2)
        row = list(tables_ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(row[0], "ai_media_task")
        self.assertEqual(row[1], "sfst")
        self.assertEqual(row[2], "天表")

    def test_json_arg_with_optional_fields(self):
        data = {
            **SAMPLE_JSON_DICT,
            "dw_layer": "ods",
            "table_format": "orc",
            "target_table_format": "hive",
            "operate_type": "新建表",
        }
        with patch(
            "sys.argv",
            ["prog", "--json", json.dumps(data, ensure_ascii=False),
             "--excel", str(self.excel_path)],
        ):
            from app.main import main
            main()

        wb = load_workbook(self.excel_path)
        row = list(wb["tables"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(row[4], "ods")
        self.assertEqual(row[7], "新建表")


class TestMainWithFile(unittest.TestCase):
    """通过 --file 参数从 JSON 文件读取的主流程集成测试。"""

    def setUp(self):
        self.tmp_dir = Path(tempfile.mkdtemp())
        self.excel_path = self.tmp_dir / "output.xlsx"
        self.json_file = self.tmp_dir / "input.json"
        self.json_file.write_text(SAMPLE_JSON_STR, encoding="utf-8")

    def tearDown(self):
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def test_file_arg_creates_excel(self):
        with patch(
            "sys.argv",
            ["prog", "--file", str(self.json_file), "--excel", str(self.excel_path)],
        ):
            from app.main import main
            main()

        self.assertTrue(self.excel_path.exists())
        wb = load_workbook(self.excel_path)
        self.assertEqual(wb["tables"].max_row, 2)
        row = list(wb["tables"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(row[0], "ai_media_task")

    def test_missing_file_exits_with_code_1(self):
        missing = self.tmp_dir / "nonexistent.json"
        with patch(
            "sys.argv",
            ["prog", "--file", str(missing), "--excel", str(self.excel_path)],
        ):
            from app.main import main
            with self.assertRaises(SystemExit) as ctx:
                main()
            self.assertEqual(ctx.exception.code, 1)


class TestMainParseFailure(unittest.TestCase):
    """输入数据解析失败时应以 exit code 1 退出。"""

    def setUp(self):
        self.tmp_dir = Path(tempfile.mkdtemp())
        self.excel_path = self.tmp_dir / "output.xlsx"

    def tearDown(self):
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def test_invalid_json_exits_with_code_1(self):
        with patch(
            "sys.argv",
            ["prog", "--json", "{bad json}", "--excel", str(self.excel_path)],
        ):
            from app.main import main
            with self.assertRaises(SystemExit) as ctx:
                main()
            self.assertEqual(ctx.exception.code, 1)

    def test_missing_required_field_exits_with_code_1(self):
        incomplete = json.dumps({"mysql_sql": SAMPLE_SQL})
        with patch(
            "sys.argv",
            ["prog", "--json", incomplete, "--excel", str(self.excel_path)],
        ):
            from app.main import main
            with self.assertRaises(SystemExit) as ctx:
                main()
            self.assertEqual(ctx.exception.code, 1)


class TestMainDateDirectory(unittest.TestCase):
    """验证默认 Excel 路径使用日期目录。"""

    def setUp(self):
        self.tmp_dir = Path(tempfile.mkdtemp())

    def tearDown(self):
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    @patch("app.main.OUTPUT_BASE_DIR")
    @patch("app.main.datetime")
    def test_default_excel_goes_to_date_directory(self, mock_dt, mock_base_dir):
        mock_dt.now.return_value.strftime.return_value = "20260318"
        mock_base_dir.__truediv__ = lambda self_, d: self.tmp_dir / d

        with patch(
            "sys.argv",
            ["prog", "--json", SAMPLE_JSON_STR],
        ):
            from app.main import main
            main()

        expected = self.tmp_dir / "20260318" / "create_table_info.xlsx"
        self.assertTrue(expected.exists())
        wb = load_workbook(expected)
        self.assertEqual(wb["tables"].max_row, 2)


class TestMainTableComment(unittest.TestCase):
    """table_comment 功能的端到端集成测试。"""

    def setUp(self):
        self.tmp_dir = Path(tempfile.mkdtemp())
        self.excel_path = self.tmp_dir / "output.xlsx"

    def tearDown(self):
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def test_table_comment_from_json_field(self):
        data = {**SAMPLE_JSON_DICT, "table_comment": "自定义表注释"}
        with patch(
            "sys.argv",
            ["prog", "--json", json.dumps(data, ensure_ascii=False),
             "--excel", str(self.excel_path)],
        ):
            from app.main import main
            main()

        wb = load_workbook(self.excel_path)
        row = list(wb["tables"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(row[3], "自定义表注释")

    def test_table_comment_fallback_from_sql(self):
        with patch(
            "sys.argv",
            ["prog", "--json", SAMPLE_JSON_STR, "--excel", str(self.excel_path)],
        ):
            from app.main import main
            main()

        wb = load_workbook(self.excel_path)
        row = list(wb["tables"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(row[3], "AI媒体任务表")

    def test_table_comment_json_overrides_sql(self):
        data = {**SAMPLE_JSON_DICT, "table_comment": "覆盖SQL注释"}
        with patch(
            "sys.argv",
            ["prog", "--json", json.dumps(data, ensure_ascii=False),
             "--excel", str(self.excel_path)],
        ):
            from app.main import main
            main()

        wb = load_workbook(self.excel_path)
        row = list(wb["tables"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(row[3], "覆盖SQL注释")


class TestMainFileWithDoubleQuotes(unittest.TestCase):
    """通过 --file 传入含未转义双引号 SQL 的 JSON 文件，验证端到端正常。"""

    def setUp(self):
        self.tmp_dir = Path(tempfile.mkdtemp())
        self.excel_path = self.tmp_dir / "output.xlsx"
        self.json_file = self.tmp_dir / "input.json"
        broken_json = (
            '{"mysql_sql": "CREATE TABLE `t` ('
            '`a` varchar(64) DEFAULT "" COMMENT "名称"'
            ') ENGINE=InnoDB COMMENT "测试表"", '
            '"day_or_hour": "天表", "product_line": "sfst"}'
        )
        self.json_file.write_text(broken_json, encoding="utf-8")

    def tearDown(self):
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def test_file_with_double_quotes_creates_excel(self):
        with patch(
            "sys.argv",
            ["prog", "--file", str(self.json_file), "--excel", str(self.excel_path)],
        ):
            from app.main import main
            main()

        self.assertTrue(self.excel_path.exists())
        wb = load_workbook(self.excel_path)
        tables_ws = wb["tables"]
        self.assertEqual(tables_ws.max_row, 2)
        row = list(tables_ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(row[0], "t")
        self.assertEqual(row[1], "sfst")
        self.assertEqual(row[2], "天表")

    def test_fields_sheet_sql_has_single_quotes_after_repair(self):
        with patch(
            "sys.argv",
            ["prog", "--file", str(self.json_file), "--excel", str(self.excel_path)],
        ):
            from app.main import main
            main()

        wb = load_workbook(self.excel_path)
        fields_row = list(
            wb["fields"].iter_rows(min_row=2, max_row=2, values_only=True)
        )[0]
        sql_in_excel = fields_row[5]
        self.assertNotIn('"', sql_in_excel)
        self.assertIn("DEFAULT ''", sql_in_excel)
        self.assertIn("COMMENT '名称'", sql_in_excel)
        self.assertIn("COMMENT '测试表'", sql_in_excel)


if __name__ == "__main__":
    unittest.main()
